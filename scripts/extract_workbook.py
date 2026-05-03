import json
import sys
from pathlib import Path
from datetime import date, datetime
from decimal import Decimal

from openpyxl import load_workbook


def cell_id(cell):
    return f"{cell.column_letter}{cell.row}"

def json_value(v):
    if isinstance(v, (datetime, date)):
        return v.isoformat()
    if isinstance(v, Decimal):
        return float(v)
    return v


def main():
    if len(sys.argv) < 3:
        print("Usage: python scripts/extract_workbook.py <input.xlsx> <output.json>")
        raise SystemExit(2)

    xlsx_path = Path(sys.argv[1])
    out_path = Path(sys.argv[2])

    wb = load_workbook(filename=xlsx_path, data_only=False, keep_links=True)

    workbook = {"sheets": {}}

    for ws in wb.worksheets:
        sheet_name = ws.title
        cells = []

        # Use worksheet dimensions; if missing, iterate through all cells with values/formulas.
        min_row = ws.min_row
        max_row = ws.max_row
        min_col = ws.min_column
        max_col = ws.max_column

        for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
            for c in row:
                if c.value is None:
                    continue

                v = c.value
                is_formula = isinstance(v, str) and v.startswith("=")

                cells.append(
                    {
                        "id": cell_id(c),
                        "nome": None,
                        "valor": None if is_formula else json_value(v),
                        "formula": v[1:] if is_formula else None,
                        "tipo": "calculado" if is_formula else "manual",
                        "unidade": None,
                        "origem": sheet_name,
                    }
                )

        workbook["sheets"][sheet_name] = cells

    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(json.dumps(workbook, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"OK: wrote {out_path} ({out_path.stat().st_size} bytes)")


if __name__ == "__main__":
    main()

